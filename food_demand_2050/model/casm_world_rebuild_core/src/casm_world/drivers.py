"""IIASA SSP v3.2 population and GDP-per-capita driver adapter.

This module only builds exogenous annual growth indices and a coverage report.
It does not run the CASM-World equilibrium model. Missing IIASA regions are
resolved through configuration-declared fallbacks; a missing value is never
interpreted as zero.
"""

from __future__ import annotations

import argparse
from collections import Counter
import json
from math import exp, isfinite, log
from pathlib import Path
import unicodedata
from typing import Any, Iterable, Iterator, Mapping, Sequence

import pandas as pd
import yaml

from casm_world.benchmark import country_codebook
from casm_world.geography import (
    EXPECTED_TERRITORIES,
    load_territory_config,
    territory_crosswalk,
)
from casm_world.paths import load_source_catalog, verify_source


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG_PATH = PROJECT_ROOT / "config" / "scenarios.yaml"
EXPECTED_SCENARIOS = ("SSP1", "SSP2", "SSP3", "SSP4", "SSP5")
EXPECTED_NODES = (2025, 2030, 2035, 2040, 2045, 2050)
PREFERRED_GDP_UNIT = "USD_2015/yr"


def _mapping(value: object, label: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise ValueError(f"{label} must be a mapping")
    return value


def _project_path(project_root: Path, relative: str, label: str) -> Path:
    path = Path(relative)
    if path.is_absolute():
        raise ValueError(f"{label} must be relative to the new project")
    root = project_root.resolve()
    resolved = (root / path).resolve()
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"{label} escapes the new project: {path}") from exc
    return resolved


def load_scenario_config(path: Path | str = DEFAULT_CONFIG_PATH) -> dict[str, Any]:
    """Load and validate the frozen SSP driver specification."""

    config_path = Path(path).resolve()
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    config = dict(_mapping(config, "scenario configuration"))
    if config.get("schema_version") != 1:
        raise ValueError("scenarios.yaml schema_version must equal 1")
    source = _mapping(config.get("source"), "source")
    series = _mapping(config.get("series"), "series")
    population = _mapping(series.get("population"), "series.population")
    gdp = _mapping(series.get("gdp_ppp_per_capita"), "series.gdp_ppp_per_capita")
    if source.get("catalog_key") != "iiasa_ssp_basic_drivers":
        raise ValueError("SSP drivers must use the verified IIASA catalog snapshot")
    if (population.get("model"), population.get("variable"), population.get("unit")) != (
        "IIASA-WiC POP 2025",
        "Population",
        "million",
    ):
        raise ValueError("Population series must be IIASA-WiC POP 2025 Population (million)")
    if (gdp.get("model"), gdp.get("variable"), gdp.get("unit")) != (
        "OECD ENV-Growth 2025",
        "GDP|PPP [per capita]",
        PREFERRED_GDP_UNIT,
    ):
        raise ValueError(
            "GDP series must be OECD ENV-Growth 2025 GDP|PPP [per capita] "
            "(USD_2015/yr)"
        )
    if tuple(config.get("scenarios", ())) != EXPECTED_SCENARIOS:
        raise ValueError("Exactly SSP1-SSP5 must be configured in order")
    if tuple(config.get("projection_nodes", ())) != EXPECTED_NODES:
        raise ValueError("Projection nodes must be 2025-2050 five-year points")
    interpolation = _mapping(config.get("interpolation"), "interpolation")
    if interpolation.get("method") != "log_linear":
        raise ValueError("Five-year SSP nodes must use log-linear interpolation")
    if interpolation.get("missing_to_zero") != "forbidden":
        raise ValueError("NA-to-zero conversion must remain forbidden")
    account = _mapping(config.get("account_universe"), "account_universe")
    if int(account.get("expected_accounts", -1)) != 193:
        raise ValueError("The scenario adapter must target exactly 193 accounts")
    if int(account.get("expected_mapped_territories", -1)) != 25:
        raise ValueError("The scenario adapter must apply all 25 territory mappings")
    return config


def log_interpolate_nodes(
    nodes: Mapping[int, float], years: Iterable[int]
) -> dict[int, float]:
    """Interpolate positive nodes linearly in logarithms.

    Extrapolation is deliberately forbidden. This makes a missing historical
    bridge or a missing 2050 node fatal instead of silently extending a path.
    """

    if not nodes:
        raise ValueError("Interpolation nodes must not be empty")
    clean: dict[int, float] = {}
    for raw_year, raw_value in nodes.items():
        year = int(raw_year)
        value = float(raw_value)
        if not isfinite(value) or value <= 0:
            raise ValueError(f"Interpolation node {year} must be finite and positive")
        clean[year] = value
    ordered = sorted(clean)
    requested = sorted({int(year) for year in years})
    if requested and (requested[0] < ordered[0] or requested[-1] > ordered[-1]):
        raise ValueError("Log interpolation does not extrapolate beyond supplied nodes")

    output: dict[int, float] = {}
    for year in requested:
        if year in clean:
            output[year] = clean[year]
            continue
        left = max(node for node in ordered if node < year)
        right = min(node for node in ordered if node > year)
        weight = (year - left) / (right - left)
        output[year] = exp((1.0 - weight) * log(clean[left]) + weight * log(clean[right]))
    return output


def _normalise_name(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value)).casefold()
    return "".join(character for character in text if character.isalnum())


def _iter_sheet_rows(workbook: object, sheet_name: str, backend: str) -> Iterator[Sequence[Any]]:
    if backend == "calamine":
        yield from workbook.get_sheet_by_name(sheet_name).iter_rows()
    else:
        yield from workbook[sheet_name].iter_rows(values_only=True)


def _open_workbook(path: Path) -> tuple[object, str]:
    """Use the optional fast reader when present, with openpyxl as fallback."""

    try:
        from python_calamine import CalamineWorkbook

        return CalamineWorkbook.from_path(str(path)), "calamine"
    except ImportError:
        from openpyxl import load_workbook

        return load_workbook(path, read_only=True, data_only=True), "openpyxl"


def _close_workbook(workbook: object, backend: str) -> None:
    if backend == "openpyxl":
        workbook.close()


def _year_column(value: object) -> int | None:
    text = str(value).strip()
    try:
        year = int(float(text))
    except ValueError:
        return None
    return year if 1900 <= year <= 2200 else None


def read_iiasa_driver_rows(workbook_path: Path, config: Mapping[str, Any]) -> dict:
    """Read only the configured population/GDP rows and required node years."""

    source = _mapping(config["source"], "source")
    series = _mapping(config["series"], "series")
    pop = _mapping(series["population"], "population")
    gdp = _mapping(series["gdp_ppp_per_capita"], "gdp")
    historical = str(config["historical_scenario"])
    scenarios = set(config["scenarios"]) | {historical}
    years = {
        int(config["interpolation"]["historical_node"]),
        int(config["interpolation"]["population_fallback_last_observed_year"]),
        *EXPECTED_NODES,
    }

    selections = {
        str(source["main_sheet"]): {
            (str(pop["model"]), str(pop["variable"]), str(pop["unit"])): "population",
            (str(gdp["model"]), str(gdp["variable"]), str(gdp["unit"])): "gdp_pc",
            (
                str(gdp["model"]),
                str(gdp["variable"]),
                str(gdp["alternate_unit"]),
            ): "gdp_pc",
        },
        str(source["turbulent_economy_sheet"]): {
            (
                str(gdp["turbulent_model"]),
                str(gdp["variable"]),
                str(gdp["unit"]),
            ): "gdp_pc",
        },
    }
    lookup: dict[tuple[str, str, str, str, str], dict[int, float]] = {}
    workbook, backend = _open_workbook(workbook_path)
    try:
        for sheet_name, accepted in selections.items():
            rows = _iter_sheet_rows(workbook, sheet_name, backend)
            try:
                header = next(rows)
            except StopIteration as exc:
                raise ValueError(f"Empty IIASA worksheet: {sheet_name}") from exc
            positions = {str(value): index for index, value in enumerate(header)}
            required = {"Model", "Scenario", "Region", "Variable", "Unit"}
            if not required <= set(positions):
                raise ValueError(f"IIASA sheet {sheet_name} lacks required columns")
            year_positions = {
                year: index
                for index, value in enumerate(header)
                if (year := _year_column(value)) in years
            }
            for row in rows:
                identity = (
                    str(row[positions["Model"]]),
                    str(row[positions["Variable"]]),
                    str(row[positions["Unit"]]),
                )
                metric = accepted.get(identity)
                scenario = str(row[positions["Scenario"]])
                if metric is None or scenario not in scenarios:
                    continue
                region = str(row[positions["Region"]]).strip()
                values: dict[int, float] = {}
                for year, position in year_positions.items():
                    raw = row[position]
                    if raw is None or raw == "":
                        continue
                    value = float(raw)
                    if not isfinite(value) or value <= 0:
                        raise ValueError(
                            f"Nonpositive IIASA value: {sheet_name}/{region}/{scenario}/{year}"
                        )
                    values[year] = value
                key = (sheet_name, metric, identity[2], scenario, region)
                if key in lookup:
                    raise ValueError(f"Duplicate IIASA driver row: {key}")
                lookup[key] = values
    finally:
        _close_workbook(workbook, backend)
    return lookup


def _nodes(
    raw: Mapping[tuple[str, str, str, str, str], Mapping[int, float]],
    *,
    sheet: str,
    metric: str,
    unit: str,
    scenario: str,
    region: str,
    years: Iterable[int],
) -> dict[int, float] | None:
    values = raw.get((sheet, metric, unit, scenario, region))
    if values is None:
        return None
    requested = tuple(int(year) for year in years)
    if not all(year in values for year in requested):
        return None
    return {year: float(values[year]) for year in requested}


def _load_accounts(project_root: Path, config: Mapping[str, Any]) -> list[str]:
    account = _mapping(config["account_universe"], "account_universe")
    path = _project_path(project_root, str(account["path"]), "account_universe.path")
    column = str(account["economy_column"])
    frame = pd.read_csv(path, usecols=[column])
    economies = sorted(frame[column].astype(str).str.strip().str.upper().unique())
    expected = int(account["expected_accounts"])
    if len(economies) != expected:
        raise ValueError(f"Expected {expected} model accounts, found {len(economies)}")
    if set(economies) & EXPECTED_TERRITORIES:
        raise ValueError("Mapped territories must not survive in the 193-account universe")
    if "OTHER_EASTERN_ASIA" not in economies or "CHN" not in economies:
        raise ValueError("Account universe must contain CHN and OTHER_EASTERN_ASIA")
    return economies


def _region_crosswalk(
    component_codes: Iterable[str],
    raw: Mapping[tuple[str, str, str, str, str], Mapping[int, float]],
    config: Mapping[str, Any],
    un_m49_path: Path,
) -> dict[str, str]:
    aliases = {str(key).upper(): str(value) for key, value in config["region_aliases"].items()}
    codebook = country_codebook(un_m49_path).set_index("economy_id")
    available_regions = {
        key[4]
        for key in raw
        if key[1] in {"population", "gdp_pc"}
    }
    normalised: dict[str, list[str]] = {}
    for region in available_regions:
        normalised.setdefault(_normalise_name(region), []).append(region)

    result: dict[str, str] = {}
    for code in sorted(set(component_codes)):
        if code in aliases:
            candidate = aliases[code]
            if candidate not in available_regions:
                raise ValueError(f"Configured IIASA alias is absent for {code}: {candidate}")
            result[code] = candidate
            continue
        if code not in codebook.index:
            raise ValueError(f"No UN/IIASA name mapping for component {code}")
        source_name = str(codebook.at[code, "economy_name"])
        matches = normalised.get(_normalise_name(source_name), [])
        if len(matches) != 1:
            raise ValueError(
                f"IIASA region match for {code}/{source_name!r} is not unique: {matches}"
            )
        result[code] = matches[0]
    return result


def _annual_path(
    history_2020: float,
    history_2025: float,
    scenario_nodes: Mapping[int, float],
) -> dict[int, float]:
    """Join history to an SSP path at a common 2025 historical anchor.

    Some GDP projections already differ slightly by SSP in 2025. Rebasing every
    path to Historical Reference 2025 keeps 2023--2024 scenario independent
    while preserving each SSP's post-2025 growth rates.
    """

    scenario_2025 = float(scenario_nodes[2025])
    scale = float(history_2025) / scenario_2025
    rebased = {int(year): float(value) * scale for year, value in scenario_nodes.items()}
    nodes = {2020: float(history_2020), **rebased}
    return log_interpolate_nodes(nodes, range(2023, 2051))


def _population_paths(
    components: Sequence[str],
    regions: Mapping[str, str],
    raw: Mapping[tuple[str, str, str, str, str], Mapping[int, float]],
    config: Mapping[str, Any],
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    main_sheet = str(config["source"]["main_sheet"])
    historical = str(config["historical_scenario"])
    pop_unit = str(config["series"]["population"]["unit"])
    proxies = {str(key): str(value) for key, value in config["population_growth_proxies"].items()}
    records: list[dict[str, Any]] = []
    coverage: list[dict[str, Any]] = []

    for code in components:
        region = regions[code]
        history = _nodes(
            raw,
            sheet=main_sheet,
            metric="population",
            unit=pop_unit,
            scenario=historical,
            region=region,
            years=(2020, 2025),
        )
        direct = all(
            _nodes(
                raw,
                sheet=main_sheet,
                metric="population",
                unit=pop_unit,
                scenario=scenario,
                region=region,
                years=EXPECTED_NODES,
            )
            is not None
            for scenario in EXPECTED_SCENARIOS
        )
        proxy_code = None if direct else proxies.get(code)
        if not direct and proxy_code is None:
            raise ValueError(f"Population fallback is not declared for {code}")
        if proxy_code is not None and proxy_code not in regions:
            raise ValueError(f"Population proxy {proxy_code} for {code} is unavailable")
        if direct and history is None:
            raise ValueError(f"Population Historical Reference 2020/2025 missing for {code}")
        fallback_last_year = int(
            config["interpolation"]["population_fallback_last_observed_year"]
        )
        if not direct:
            own_last = _nodes(
                raw,
                sheet=main_sheet,
                metric="population",
                unit=pop_unit,
                scenario=historical,
                region=region,
                years=(fallback_last_year,),
            )
            proxy_region = regions[str(proxy_code)]
            proxy_history = _nodes(
                raw,
                sheet=main_sheet,
                metric="population",
                unit=pop_unit,
                scenario=historical,
                region=proxy_region,
                years=(fallback_last_year, 2020, 2025),
            )
            if own_last is None or proxy_history is None:
                raise ValueError(
                    f"Population fallback bridge lacks {fallback_last_year}/2020 data for {code}"
                )
        coverage.append(
            {
                "economy_id": code,
                "region": region,
                "status": "direct_iiasa" if direct else "fallback_proxy_growth",
                "proxy_economy_id": proxy_code,
                "fallback_anchor_year": None if direct else fallback_last_year,
            }
        )

        for scenario in EXPECTED_SCENARIOS:
            if direct:
                nodes = _nodes(
                    raw,
                    sheet=main_sheet,
                    metric="population",
                    unit=pop_unit,
                    scenario=scenario,
                    region=region,
                    years=EXPECTED_NODES,
                )
            else:
                proxy_region = regions[str(proxy_code)]
                proxy_nodes = _nodes(
                    raw,
                    sheet=main_sheet,
                    metric="population",
                    unit=pop_unit,
                    scenario=scenario,
                    region=proxy_region,
                    years=EXPECTED_NODES,
                )
                if proxy_nodes is None:
                    raise ValueError(f"Population proxy path missing for {code}: {proxy_code}")
                own_anchor = own_last[fallback_last_year]
                proxy_anchor = proxy_history[fallback_last_year]
                nodes = {
                    year: own_anchor * value / proxy_anchor
                    for year, value in proxy_nodes.items()
                }
            assert nodes is not None
            history_2020 = (
                history[2020]
                if direct
                else own_last[fallback_last_year]
                * proxy_history[2020]
                / proxy_history[fallback_last_year]
            )
            history_2025 = (
                history[2025]
                if direct
                else own_last[fallback_last_year]
                * proxy_history[2025]
                / proxy_history[fallback_last_year]
            )
            annual = _annual_path(history_2020, history_2025, nodes)
            records.extend(
                {
                    "economy_id": code,
                    "scenario": scenario,
                    "year": year,
                    "population_million": value,
                }
                for year, value in annual.items()
            )
    return pd.DataFrame.from_records(records), coverage


def _gdp_conversion_factor(
    raw: Mapping[tuple[str, str, str, str, str], Mapping[int, float]],
    config: Mapping[str, Any],
) -> float:
    gdp = config["series"]["gdp_ppp_per_capita"]
    rule = gdp["alternate_unit_conversion"]
    sheet = str(config["source"]["main_sheet"])
    region = str(rule["reference_region"])
    scenario = str(rule["reference_scenario"])
    year = int(rule["reference_year"])
    preferred = _nodes(
        raw,
        sheet=sheet,
        metric="gdp_pc",
        unit=str(gdp["unit"]),
        scenario=scenario,
        region=region,
        years=(year,),
    )
    alternate = _nodes(
        raw,
        sheet=sheet,
        metric="gdp_pc",
        unit=str(gdp["alternate_unit"]),
        scenario=scenario,
        region=region,
        years=(year,),
    )
    if preferred is None or alternate is None:
        raise ValueError("World GDP unit conversion reference is missing")
    factor = preferred[year] / alternate[year]
    if not isfinite(factor) or factor <= 0:
        raise ValueError("GDP alternate-unit conversion factor is invalid")
    return factor


def _gdp_paths(
    contributing_components: Sequence[str],
    scope_excluded: Sequence[str],
    regions: Mapping[str, str],
    raw: Mapping[tuple[str, str, str, str, str], Mapping[int, float]],
    config: Mapping[str, Any],
) -> tuple[pd.DataFrame, list[dict[str, Any]], float]:
    main_sheet = str(config["source"]["main_sheet"])
    turbulent_sheet = str(config["source"]["turbulent_economy_sheet"])
    historical = str(config["historical_scenario"])
    gdp = config["series"]["gdp_ppp_per_capita"]
    preferred_unit = str(gdp["unit"])
    alternate_unit = str(gdp["alternate_unit"])
    proxies = {str(key): str(value) for key, value in config["gdp_growth_and_level_proxies"].items()}
    conversion = _gdp_conversion_factor(raw, config)
    records: list[dict[str, Any]] = []
    coverage: list[dict[str, Any]] = [
        {
            "economy_id": code,
            "region": regions[code],
            "status": "gdp_scope_excluded_parent_inclusive",
            "proxy_economy_id": None,
            "source_unit": None,
        }
        for code in sorted(scope_excluded)
    ]

    for code in contributing_components:
        region = regions[code]

        def complete(sheet: str, unit: str) -> bool:
            return all(
                _nodes(
                    raw,
                    sheet=sheet,
                    metric="gdp_pc",
                    unit=unit,
                    scenario=scenario,
                    region=region,
                    years=EXPECTED_NODES,
                )
                is not None
                for scenario in EXPECTED_SCENARIOS
            )

        if complete(main_sheet, preferred_unit):
            route = "direct_iiasa"
            route_sheet = main_sheet
            route_unit = preferred_unit
            factor = 1.0
            proxy_code = None
        elif complete(turbulent_sheet, preferred_unit):
            route = "official_turbulent_supplement"
            route_sheet = turbulent_sheet
            route_unit = preferred_unit
            factor = 1.0
            proxy_code = None
        elif complete(main_sheet, alternate_unit):
            route = "fallback_alternate_unit_rescaled"
            route_sheet = main_sheet
            route_unit = alternate_unit
            factor = conversion
            proxy_code = None
        else:
            route = "fallback_proxy_growth_and_level"
            route_sheet = main_sheet
            route_unit = preferred_unit
            factor = 1.0
            proxy_code = proxies.get(code)
            if proxy_code is None:
                raise ValueError(f"GDP fallback is not declared for {code}")
            proxy_region = regions.get(proxy_code)
            if proxy_region is None:
                raise ValueError(f"GDP proxy {proxy_code} for {code} is unavailable")
            if not all(
                _nodes(
                    raw,
                    sheet=main_sheet,
                    metric="gdp_pc",
                    unit=preferred_unit,
                    scenario=scenario,
                    region=proxy_region,
                    years=EXPECTED_NODES,
                )
                is not None
                for scenario in EXPECTED_SCENARIOS
            ):
                raise ValueError(f"GDP proxy lacks a complete preferred path: {proxy_code}")

        coverage.append(
            {
                "economy_id": code,
                "region": region,
                "status": route,
                "proxy_economy_id": proxy_code,
                "source_unit": route_unit,
            }
        )

        if proxy_code is None:
            history = _nodes(
                raw,
                sheet=main_sheet,
                metric="gdp_pc",
                unit=route_unit,
                scenario=historical,
                region=region,
                years=(2020, 2025),
            )
            if history is None:
                raise ValueError(f"GDP Historical Reference missing for {code}/{route_unit}")
            history = {year: value * factor for year, value in history.items()}
        else:
            proxy_region = regions[proxy_code]
            history = _nodes(
                raw,
                sheet=main_sheet,
                metric="gdp_pc",
                unit=preferred_unit,
                scenario=historical,
                region=proxy_region,
                years=(2020, 2025),
            )
            if history is None:
                raise ValueError(f"GDP proxy history missing for {proxy_code}")

        for scenario in EXPECTED_SCENARIOS:
            source_region = regions[proxy_code] if proxy_code else region
            nodes = _nodes(
                raw,
                sheet=route_sheet,
                metric="gdp_pc",
                unit=route_unit,
                scenario=scenario,
                region=source_region,
                years=EXPECTED_NODES,
            )
            if nodes is None:
                raise ValueError(f"GDP scenario nodes missing for {code}/{scenario}")
            nodes = {year: value * factor for year, value in nodes.items()}
            annual = _annual_path(history[2020], history[2025], nodes)
            records.extend(
                {
                    "economy_id": code,
                    "scenario": scenario,
                    "year": year,
                    "gdp_pc_usd_2015": value,
                }
                for year, value in annual.items()
            )
    return pd.DataFrame.from_records(records), coverage, conversion


def _status_counts(records: Sequence[Mapping[str, Any]]) -> dict[str, int]:
    return dict(sorted(Counter(str(record["status"]) for record in records).items()))


def build_ssp_drivers(
    project_root: Path | str = PROJECT_ROOT,
    config_path: Path | str | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Build annual 2023-2050 indices for the actual 193 model accounts."""

    root = Path(project_root).resolve()
    scenario_path = Path(config_path).resolve() if config_path else root / "config/scenarios.yaml"
    config = load_scenario_config(scenario_path)
    accounts = _load_accounts(root, config)
    territory_path = _project_path(
        root,
        str(config["account_universe"]["territory_config"]),
        "account_universe.territory_config",
    )
    territory_config = load_territory_config(territory_path)
    crosswalk = territory_crosswalk(territory_config)
    if len(crosswalk) != 25:
        raise ValueError("All 25 territory mappings must be applied")

    named_target = str(config["account_universe"]["named_target_without_direct_component"])
    if named_target != "OTHER_EASTERN_ASIA":
        raise ValueError("The TWN receiver must remain OTHER_EASTERN_ASIA")
    base_components = sorted(set(accounts) - {named_target})
    components = sorted(set(base_components) | set(crosswalk["source_economy_id"]))
    if "TWN" not in components or "CHN" not in components:
        raise ValueError("TWN and CHN components must both be explicit")

    catalog = load_source_catalog(root / "config/data_sources.yaml")
    iiasa = catalog.source(str(config["source"]["catalog_key"]))
    un_m49 = catalog.source("un_m49")
    verify_source(iiasa)
    verify_source(un_m49)
    raw = read_iiasa_driver_rows(iiasa.path, config)
    regions = _region_crosswalk(components, raw, config, un_m49.path)

    population, population_coverage = _population_paths(
        components, regions, raw, config
    )
    scope_excluded = sorted(
        crosswalk.loc[
            crosswalk["gdp_scope_action"].eq("already_in_accounting_target"),
            "source_economy_id",
        ]
    )
    gdp_components = sorted(set(components) - set(scope_excluded))
    gdp_pc, gdp_coverage, conversion = _gdp_paths(
        gdp_components, scope_excluded, regions, raw, config
    )

    source_to_target = crosswalk.set_index("source_economy_id")["accounting_target"].to_dict()
    population["accounting_target"] = population["economy_id"].map(source_to_target).fillna(
        population["economy_id"]
    )
    account_population = (
        population.groupby(["accounting_target", "scenario", "year"], as_index=False)[
            "population_million"
        ]
        .sum()
        .rename(columns={"accounting_target": "economy_id"})
    )

    gdp_work = gdp_pc.merge(
        population[["economy_id", "scenario", "year", "population_million"]],
        on=["economy_id", "scenario", "year"],
        how="left",
        validate="one_to_one",
    )
    if gdp_work["population_million"].isna().any():
        raise ValueError("GDP aggregation lacks component population weights")
    gdp_work["gdp_billion_2015"] = (
        gdp_work["gdp_pc_usd_2015"] * gdp_work["population_million"] / 1000.0
    )
    gdp_work["accounting_target"] = gdp_work["economy_id"].map(source_to_target).fillna(
        gdp_work["economy_id"]
    )
    account_gdp = gdp_work.groupby(
        ["accounting_target", "scenario", "year"], as_index=False
    )["gdp_billion_2015"].sum()
    combined = account_population.merge(
        account_gdp,
        left_on=["economy_id", "scenario", "year"],
        right_on=["accounting_target", "scenario", "year"],
        how="left",
        validate="one_to_one",
    ).drop(columns="accounting_target")
    combined["gdp_pc_usd_2015"] = (
        combined["gdp_billion_2015"] * 1000.0 / combined["population_million"]
    )

    actual_accounts = set(combined["economy_id"])
    if actual_accounts != set(accounts):
        raise ValueError(
            "SSP aggregation does not match the model accounts: "
            f"missing={sorted(set(accounts)-actual_accounts)}, "
            f"extra={sorted(actual_accounts-set(accounts))}"
        )
    expected_rows = len(accounts) * len(EXPECTED_SCENARIOS) * 28
    if len(combined) != expected_rows:
        raise ValueError(f"Expected {expected_rows} complete driver rows, found {len(combined)}")
    numeric = combined[["population_million", "gdp_pc_usd_2015"]]
    if numeric.isna().any().any() or not numeric.map(lambda value: isfinite(float(value))).all().all():
        raise ValueError("SSP drivers contain missing or non-finite values")
    if numeric.le(0).any().any():
        raise ValueError("SSP drivers contain nonpositive values; zero fill is forbidden")

    anchors = combined.loc[
        combined["year"].eq(2025),
        ["economy_id", "scenario", "population_million", "gdp_pc_usd_2015"],
    ].rename(
        columns={
            "population_million": "population_2025",
            "gdp_pc_usd_2015": "gdp_pc_2025",
        }
    )
    combined = combined.merge(
        anchors, on=["economy_id", "scenario"], how="left", validate="many_to_one"
    )
    combined["population_index_2025"] = (
        combined["population_million"] / combined["population_2025"]
    )
    combined["gdp_ppp_per_capita_index_2025"] = (
        combined["gdp_pc_usd_2015"] / combined["gdp_pc_2025"]
    )
    drivers = combined[
        [
            "economy_id",
            "scenario",
            "year",
            "population_million",
            "gdp_billion_2015",
            "gdp_pc_usd_2015",
            "population_index_2025",
            "gdp_ppp_per_capita_index_2025",
        ]
    ].sort_values(["economy_id", "scenario", "year"]).reset_index(drop=True)
    anchor_values = drivers.loc[
        drivers["year"].eq(2025),
        ["population_index_2025", "gdp_ppp_per_capita_index_2025"],
    ]
    if not anchor_values.eq(1.0).all().all():
        raise AssertionError("Every SSP index must equal one in 2025")

    population_fallbacks = [
        record for record in population_coverage if record["status"] != "direct_iiasa"
    ]
    gdp_non_direct = [
        record for record in gdp_coverage if record["status"] != "direct_iiasa"
    ]
    expected_routes = config["expected_coverage_routes"]
    actual_supplement = {
        row["economy_id"]
        for row in gdp_coverage
        if row["status"] == "official_turbulent_supplement"
    }
    actual_alternate = {
        row["economy_id"]
        for row in gdp_coverage
        if row["status"] == "fallback_alternate_unit_rescaled"
    }
    if actual_supplement != set(expected_routes["official_turbulent_supplement"]):
        raise ValueError("Unexpected official turbulent-economy coverage set")
    if actual_alternate != set(expected_routes["alternate_gdp_unit_rescaled"]):
        raise ValueError("Unexpected alternate-unit GDP fallback set")
    if set(scope_excluded) != set(expected_routes["gdp_scope_excluded_parent_inclusive"]):
        raise ValueError("Unexpected parent-inclusive GDP scope set")

    fallback_components = {
        record["economy_id"]
        for record in [*population_fallbacks, *gdp_non_direct]
    }
    fallback_accounts = sorted({source_to_target.get(code, code) for code in fallback_components})
    report: dict[str, Any] = {
        "source_id": iiasa.source_id,
        "population_series": {
            "model": config["series"]["population"]["model"],
            "variable": config["series"]["population"]["variable"],
            "unit": config["series"]["population"]["unit"],
        },
        "gdp_series": {
            "model": config["series"]["gdp_ppp_per_capita"]["model"],
            "variable": config["series"]["gdp_ppp_per_capita"]["variable"],
            "unit": config["series"]["gdp_ppp_per_capita"]["unit"],
        },
        "scenarios": list(EXPECTED_SCENARIOS),
        "projection_nodes": list(EXPECTED_NODES),
        "output_year_start": 2023,
        "output_year_end": 2050,
        "output_units": {
            "population_million": "million persons",
            "gdp_billion_2015": "billion 2015 USD PPP",
            "gdp_pc_usd_2015": "2015 USD PPP per person",
            "indices": "2025=1",
        },
        "anchor_year": 2025,
        "interpolation": "log_linear",
        "pre_anchor_method": config["interpolation"]["pre_anchor_method"],
        "population_fallback_bridge": "declared_proxy_growth_from_last_observed_2015",
        "model_account_count": len(accounts),
        "source_component_count": len(components),
        "gdp_contributing_component_count": len(gdp_components),
        "territory_mappings_applied": len(crosswalk),
        "driver_row_count": len(drivers),
        "population_status_counts": _status_counts(population_coverage),
        "gdp_status_counts": _status_counts(gdp_coverage),
        "population_fallbacks": population_fallbacks,
        "gdp_non_direct_routes": gdp_non_direct,
        "accounts_affected_by_any_fallback": fallback_accounts,
        "alternate_gdp_unit_to_2015_factor": conversion,
        "twn_accounting_target": source_to_target["TWN"],
        "other_eastern_asia_source_components": ["TWN"],
        "chn_contains_twn": False,
        "missing_value_count": 0,
        "zero_fill_count": 0,
        "status": "passed",
    }
    return drivers, report


def write_ssp_drivers(
    project_root: Path | str = PROJECT_ROOT,
    config_path: Path | str | None = None,
) -> tuple[Path, Path, dict[str, Any]]:
    """Generate the driver CSV and coverage JSON without running the model."""

    root = Path(project_root).resolve()
    scenario_path = Path(config_path).resolve() if config_path else root / "config/scenarios.yaml"
    config = load_scenario_config(scenario_path)
    drivers, report = build_ssp_drivers(root, scenario_path)
    driver_path = _project_path(root, str(config["outputs"]["drivers"]), "outputs.drivers")
    report_path = _project_path(
        root, str(config["outputs"]["coverage_report"]), "outputs.coverage_report"
    )
    driver_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    drivers.to_csv(driver_path, index=False)
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return driver_path, report_path, report


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT)
    parser.add_argument("--config", type=Path)
    args = parser.parse_args()
    driver_path, report_path, report = write_ssp_drivers(
        args.project_root, args.config
    )
    print(
        json.dumps(
            {
                "drivers": str(driver_path),
                "coverage_report": str(report_path),
                "model_accounts": report["model_account_count"],
                "rows": report["driver_row_count"],
                "status": report["status"],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
