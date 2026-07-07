"""Validate the Python reproduction against the GAMS ground truth.

Compares results/results_long.csv with the GAMS result workbooks
预测结果整理/3RESULTCOM-{normal,diet,median}.XLSX (all SIMs, TSP years)
and with the manuscript tables (docs/manuscript_tables/table2/4/6.csv).

Writes results/validation_report.md.

Usage:  python3 validate.py
"""

import csv
import os
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
RESULTS = os.path.join(HERE, "..", "results")
DOCS = os.path.join(HERE, "..", "docs", "manuscript_tables")
TRUTH_DIR = "/root/data/Paper/食物预测2050/预测结果整理"

TRUTH_FILES = {"PTS": "3RESULTCOM-normal.XLSX",
               "HDS": "3RESULTCOM-diet.XLSX",
               "MTS": "3RESULTCOM-median.XLSX"}
TSP = ["2023", "2024", "2035", "2050"]
SIMS = ["BASE", "SIM1", "SIM2", "SIM3", "SIM4", "SIM5", "SIM6"]
PREFIX = {"PTS": "A", "HDS": "B", "MTS": "C"}
SIM_TO_CODE = {"SIM1": 1, "SIM4": 2, "SIM5": 3, "SIM2": 4, "SIM3": 5,
               "SIM6": 6}

# GAMS sheet -> (python variable in results_long, scale python->GAMS)
VAR_SHEETS = {
    "PRODX": ("production", 1.0),
    "AREAX": ("area", 1.0),
    "YILDX": ("yield", 1.0),
    "NIMPX": ("net_import", 1.0),
    "NEXPX": ("net_export", 1.0),
    "FOODX": ("food_demand_total", 1.0),
    "PFODX": ("food_demand_pc", 1.0),
    "FEEDX": ("feed_demand", 1.0),
    "PROCX": ("processing_demand", 1.0),
    "CRSHX": ("crush_demand", 1.0),
    "SEEDX": ("seed_demand", 1.0),
    "WASTX": ("waste_demand", 1.0),
    "OTHEX": ("other_demand", 1.0),
    "STOCX": ("stock_change", 1.0),
    "DOMDX": ("domestic_demand_total", 1.0),
    "PRICX": ("consumer_price", 1.0),
}
MACRO_ROWS = {  # (sheet, row label) -> python (variable, commodity)
    ("POPUX", "TOTAL POPULATION"): ("population_total", "ALL"),
    ("POPUX", "URBAN POPULATION"): ("population_urban", "ALL"),
    ("PGDPX", "PER CAPITA GDP"): ("gdp_per_capita", "ALL"),
    ("CO2EX", "CROP EMISSION"): ("co2_crop", "ALL"),
    ("CO2EX", "LIVESTOCK EMISSION"): ("co2_livestock", "ALL"),
    ("CO2EX", "TOTAL EMISSION"): ("co2_total", "ALL"),
    ("ENGYX", "ENERGY"): ("energy_pc_day_total", "ALL"),
    ("ENGYX", "PROTEIN"): ("protein_pc_day_total", "ALL"),
    ("ENGYX", "FAT"): ("fat_pc_day_total", "ALL"),
    ("ENGYX", "CARBOHYDRATE"): ("carbohydrate_pc_day_total", "ALL"),
}
KEY_INDICATORS = [  # rows for the per-scenario key table in the report
    ("food_demand_pc", "RICE", "per-capita rice demand 2050 (kg)"),
    ("food_demand_pc", "PIGM", "per-capita pork demand 2050 (kg)"),
    ("food_demand_pc", "MILK", "per-capita dairy demand 2050 (kg)"),
    ("energy_pc_day_total", "ALL", "dietary energy 2050 (kcal/cap/day)"),
    ("co2_total", "ALL", "total CO2 2050 (万吨 CO2e)"),
]


def load_python():
    py = {}
    with open(os.path.join(RESULTS, "results_long.csv"), encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            py[(row["scenario"], row["variable"], row["commodity"],
                row["year"])] = float(row["value"])
    return py


def parse_sheet(ws):
    """-> {(row_label, sim, year): value} using the GDXXRW header layout."""
    grid = list(ws.iter_rows(values_only=True))
    # locate the SIM header row and the year row below it
    simrow = yearrow = None
    for i, row in enumerate(grid):
        vals = {str(v).strip().upper() for v in row if v is not None}
        if vals & set(SIMS):
            simrow, yearrow = i, i + 1
            break
    if simrow is None:
        return {}
    cols = {}  # col -> (sim, year)
    for j, v in enumerate(grid[simrow]):
        if v is None:
            continue
        s = str(v).strip().upper()
        if s in SIMS and j < len(grid[yearrow]):
            y = grid[yearrow][j]
            if y is not None:
                y = str(y).strip()
                y = y[:-2] if y.endswith(".0") else y
                cols[j] = (s, y)
    first_data = min(cols)
    out = {}
    for row in grid[yearrow + 1:]:
        # row label: rightmost text cell left of the data block
        label = None
        for j in range(first_data - 1, -1, -1):
            if j < len(row) and isinstance(row[j], str) and row[j].strip():
                label = row[j].strip().upper()
                break
        if label is None:
            continue
        for j, (s, y) in cols.items():
            if j < len(row) and isinstance(row[j], (int, float)):
                out[(label, s, y)] = float(row[j])
    return out


def code_of(group, sim):
    return "BS" if sim == "BASE" else f"{PREFIX[group]}{SIM_TO_CODE[sim]}"


def main():
    py = load_python()
    lines = []
    lines.append("# Validation report — Python CASM vs GAMS (3RESULTCOM)\n")
    lines.append("Python port of CASM v2.2.7 (base 2024, projections "
                 "2025-2050), 19 scenarios (BS + A1-A6 + B1-B6 + C1-C6).\n")
    lines.append("GAMS truth: 预测结果整理/3RESULTCOM-{normal,diet,median}"
                 ".XLSX. Comparison over the reported periods "
                 "TSP = 2023/2024/2035/2050.\n")

    summary_rows = []      # per scenario aggregate deviation stats
    key_rows = []          # per scenario key-indicator table
    worst = []             # worst individual cells

    for group, fname in TRUTH_FILES.items():
        wb = openpyxl.load_workbook(os.path.join(TRUTH_DIR, fname),
                                    read_only=True, data_only=True)
        sheets = {}
        for sh in list(VAR_SHEETS) + sorted({s for s, _ in MACRO_ROWS}):
            if sh in wb.sheetnames:
                sheets[sh] = parse_sheet(wb[sh])
        wb.close()

        for sim in SIMS:
            if group != "PTS" and sim == "BASE":
                continue  # BASE identical across folders; PTS covers it
            code = code_of(group, sim)
            diffs = []   # (reldiff, absdiff, gams, desc)
            ncmp = nzero = 0
            for sh, (var, scale) in VAR_SHEETS.items():
                tab = sheets.get(sh, {})
                for (label, s, y), gv in tab.items():
                    if s != sim or y not in TSP:
                        continue
                    pv = py.get((code, var, label, y))
                    if pv is None:
                        continue
                    pv *= scale
                    ncmp += 1
                    denom = max(abs(gv), 1e-9)
                    rd = abs(pv - gv) / denom
                    diffs.append((rd, abs(pv - gv), gv,
                                  f"{sh}/{label}/{y}"))
            for (sh, label), (var, com) in MACRO_ROWS.items():
                tab = sheets.get(sh, {})
                for y in TSP:
                    gv = tab.get((label, sim, y))
                    pv = py.get((code, var, com, y))
                    if gv is None or pv is None:
                        continue
                    ncmp += 1
                    rd = abs(pv - gv) / max(abs(gv), 1e-9)
                    diffs.append((rd, abs(pv - gv), gv,
                                  f"{sh}/{label}/{y}"))
            if not diffs:
                continue
            rds = sorted(d[0] for d in diffs)
            mx = max(diffs, key=lambda d: d[0])
            # ignore tiny-denominator cells for the headline max
            big = [d for d in diffs if abs(d[2]) > 1.0]
            mxb = max(big, key=lambda d: d[0]) if big else mx
            summary_rows.append(
                (code, group, sim, ncmp,
                 rds[len(rds) // 2], rds[-1], mx[3],
                 mxb[0], mxb[3]))
            worst += [(code,) + d for d in sorted(diffs, reverse=True)[:3]]

            row = [code]
            for var, com, _ in KEY_INDICATORS:
                pv = py.get((code, var, com, "2050"))
                # GAMS counterpart
                gv = None
                if var == "food_demand_pc":
                    gv = sheets.get("PFODX", {}).get((com, sim, "2050"))
                elif var == "energy_pc_day_total":
                    gv = sheets.get("ENGYX", {}).get(("ENERGY", sim, "2050"))
                elif var == "co2_total":
                    gv = sheets.get("CO2EX", {}).get(("TOTAL EMISSION", sim,
                                                      "2050"))
                row.append((pv, gv))
            key_rows.append(row)

    # ---- write report --------------------------------------------------
    lines.append("## Per-scenario deviation summary\n")
    lines.append("`n` = number of compared cells (16 variable sheets x "
                 "commodities x TSP years + 10 macro rows). Relative "
                 "deviation = |py - gams| / max(|gams|, 1e-9).\n")
    lines.append("| scenario | group | SIM | n | median rel.dev | max rel.dev"
                 " | worst cell | max rel.dev (GAMS>1) | worst cell (GAMS>1) |")
    lines.append("|---|---|---|---|---|---|---|---|---|")
    for r in summary_rows:
        lines.append(f"| {r[0]} | {r[1]} | {r[2]} | {r[3]} | {r[4]:.2e} | "
                     f"{r[5]:.2e} | {r[6]} | {r[7]:.2e} | {r[8]} |")

    lines.append("\n## Key indicators, 2050 (Python vs GAMS)\n")
    hdr = "| scenario | " + " | ".join(d for _, _, d in KEY_INDICATORS) + " |"
    lines.append(hdr)
    lines.append("|" + "---|" * (len(KEY_INDICATORS) + 1))
    for row in key_rows:
        cells = []
        for pv, gv in row[1:]:
            if pv is None:
                cells.append("-")
            elif gv is None:
                cells.append(f"{pv:.2f} / n.a.")
            else:
                cells.append(f"{pv:.2f} / {gv:.2f}")
        lines.append(f"| {row[0]} | " + " | ".join(cells) + " |")
    lines.append("\n(Each cell: Python / GAMS.)\n")

    # ---- manuscript cross-check ----------------------------------------
    lines.append("## Manuscript table cross-check\n")
    man = []
    # table 2: per-capita food demand
    t2map = {"Rice": ["RICE"], "Wheat": ["WHEA"],
             "Edible oils": ["SOYO", "RAPO", "GRDO"], "Fruits": ["FRTO"],
             "Vegetables": ["VEGT"], "Pork": ["PIGM"], "Beef": ["CATM"],
             "Mutton": ["SHGM"], "Poultry": ["CHKM"], "Eggs": ["EGGS"],
             "Dairy products": ["MILK"], "Aquatic products": ["FISH"]}
    with open(os.path.join(DOCS, "table2.csv"), encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    hdr, data = rows[0], rows[1:]
    colmap = {"2024": ("BS", "2024"), "BS 2035": ("BS", "2035"),
              "BS 2050": ("BS", "2050"), "PTS(A1) 2035": ("A1", "2035"),
              "PTS(A1) 2050": ("A1", "2050"), "HDS(B1) 2035": ("B1", "2035"),
              "HDS(B1) 2050": ("B1", "2050"), "MTS(C1) 2035": ("C1", "2035"),
              "MTS(C1) 2050": ("C1", "2050")}
    for r in data:
        food = r[0]
        if food not in t2map:
            continue
        for j, col in enumerate(hdr[1:], 1):
            if col not in colmap:
                continue
            code, y = colmap[col]
            pv = sum(py.get((code, "food_demand_pc", c, y), 0.0)
                     for c in t2map[food])
            man.append(("T2 " + food + " " + col, float(r[j]), pv))
    # table 6: CO2 (百万吨)
    with open(os.path.join(DOCS, "table6.csv"), encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    t6code = {"BS": "BS", "PTS (A1)": "A1", "HDS (B1)": "B1",
              "MTS (C1)": "C1"}
    for r in rows[1:]:
        if r[0] not in t6code:
            continue
        code = t6code[r[0]]
        for j, var in ((1, "co2_crop"), (2, "co2_livestock"), (3, "co2_total")):
            pv = py.get((code, var, "ALL", "2050"), 0.0) / 100.0
            man.append((f"T6 {r[0]} {var}", float(r[j]), pv))

    lines.append("| item | manuscript | python | rel.dev |")
    lines.append("|---|---|---|---|")
    mworst = 0.0
    for name, mv, pv in man:
        rd = abs(pv - mv) / max(abs(mv), 1e-9)
        mworst = max(mworst, rd)
        lines.append(f"| {name} | {mv:.2f} | {pv:.2f} | {rd:.1e} |")
    lines.append(f"\nWorst manuscript-table deviation: {mworst:.2e} "
                 "(manuscript values are rounded to 2 decimals).\n")

    out = os.path.join(RESULTS, "validation_report.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print("wrote", out)
    print("scenarios compared:", len(summary_rows))
    if summary_rows:
        print("max rel.dev (GAMS>1 cells):",
              max(r[7] for r in summary_rows))


if __name__ == "__main__":
    main()
