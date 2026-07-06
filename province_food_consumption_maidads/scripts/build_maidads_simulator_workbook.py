from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[2]
RESULTS = ROOT / "ProvinceMAIDADS" / "Results"
DATA_OUT = ROOT / "ProvinceMAIDADS" / "Data" / "output"
OUTPUT = RESULTS / "省级MAIDADS_Simulator.xlsx"

GROUPS = ["grain", "oil", "vegfruit", "meatsea", "dairyegg", "nonfood"]
GROUP_LABELS = {
    "grain": "Grain / staples",
    "oil": "Oils and fats",
    "vegfruit": "Vegetables and fruits",
    "meatsea": "Meat and aquatic products",
    "dairyegg": "Dairy and eggs",
    "nonfood": "Other / non-covered",
}
PRICE_COLS = {g: f"p_{g}_model" for g in GROUPS}
OBS_X_COLS = {g: f"x_{g}" for g in GROUPS}
PROJ_X_COLS = {g: f"xhat_{g}" for g in GROUPS}


def style_title(cell, fill="1F4E78", color="FFFFFF", size=16):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=color, bold=True, size=size)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(row_cells, fill="D9EAF7"):
    for cell in row_cells:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="9EADBD"))


def style_table(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")
            if cell.row > min_row and cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")


def write_df(ws, df: pd.DataFrame, start_row=1, start_col=1, header=True):
    row = start_row
    if header:
        for c, col in enumerate(df.columns, start_col):
            ws.cell(row=row, column=c, value=col)
        style_header(ws[row])
        row += 1
    for r in df.itertuples(index=False):
        for c, value in enumerate(r, start_col):
            if pd.isna(value):
                value = None
            ws.cell(row=row, column=c, value=value)
        row += 1
    return row - 1


def build_province_data() -> pd.DataFrame:
    panel = pd.read_csv(DATA_OUT / "maidads6_panel.csv")
    proj = pd.read_csv(RESULTS / "projection_province_path.csv")

    observed_rows = []
    for _, row in panel.iterrows():
        item = {
            "key": f"{row['provincechn']}|{int(row['year'])}",
            "source": "observed",
            "provincechn": row["provincechn"],
            "province": int(row["province"]),
            "year": int(row["year"]),
            "population_10k": row["population_10k"],
            "m": row["m"],
        }
        for g in GROUPS:
            item[f"p_{g}"] = row[PRICE_COLS[g]]
            item[f"x_{g}"] = row[OBS_X_COLS[g]]
        observed_rows.append(item)

    projection_rows = []
    for _, row in proj.iterrows():
        item = {
            "key": f"{row['provincechn']}|{int(row['year'])}",
            "source": "projection",
            "provincechn": row["provincechn"],
            "province": int(row["province"]),
            "year": int(row["year"]),
            "population_10k": row["population_10k"],
            "m": row["m"],
        }
        for g in GROUPS:
            item[f"p_{g}"] = row[PRICE_COLS[g]]
            item[f"x_{g}"] = row[PROJ_X_COLS[g]]
        projection_rows.append(item)

    out = pd.DataFrame(observed_rows + projection_rows)

    national_rows = []
    for year, tmp in out.groupby("year"):
        weights = tmp["population_10k"].to_numpy(float)
        item = {
            "key": f"全国加权|{int(year)}",
            "source": "national_weighted",
            "provincechn": "全国加权",
            "province": 0,
            "year": int(year),
            "population_10k": tmp["population_10k"].sum(),
            "m": float(np.average(tmp["m"], weights=weights)),
        }
        for g in GROUPS:
            item[f"p_{g}"] = float(np.average(tmp[f"p_{g}"], weights=weights))
            item[f"x_{g}"] = float(np.average(tmp[f"x_{g}"], weights=weights))
        national_rows.append(item)
    out = pd.concat([pd.DataFrame(national_rows), out], ignore_index=True)
    out = out.sort_values(["province", "year", "provincechn"]).reset_index(drop=True)
    return out


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    default = wb.active
    wb.remove(default)

    ws_readme = wb.create_sheet("Read Me")
    ws_model = wb.create_sheet("MAIDADS MODEL")
    ws_run = wb.create_sheet("RUN")
    ws_data = wb.create_sheet("Province Data")
    ws_proj = wb.create_sheet("Projection Data")
    ws_elast = wb.create_sheet("Elasticities")
    ws_checks = wb.create_sheet("Checks")
    ws_solver = wb.create_sheet("Utility Solver")
    ws_lists = wb.create_sheet("Lists")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    params = pd.read_csv(RESULTS / "parameter_estimates.csv")
    params = params[params["model"].eq("MAIDADS_sat")].set_index("group").loc[GROUPS].reset_index()
    province_data = build_province_data()
    projection_group = pd.read_csv(RESULTS / "projection_group_2030_2035_2050.csv")
    projection_feed = pd.read_csv(RESULTS / "projection_item_feed_2030_2035_2050.csv")
    income_elast = pd.read_csv(RESULTS / "elasticity_income_grid.csv")
    price_elast = pd.read_csv(RESULTS / "elasticity_price_marshallian_grid.csv")
    consistency = pd.read_csv(RESULTS / "elasticity_consistency_tests.csv")

    # Read Me
    ws_readme.merge_cells("A1:H1")
    ws_readme["A1"] = "China Provincial MAIDADS Simulator"
    style_title(ws_readme["A1"])
    readme_rows = [
        ("Purpose", "Formula-driven simulator based on the latest provincial MAIDADS estimates."),
        ("Template", "Mimics the original structure: Read Me, model equations, RUN, data and checks."),
        ("No macros", "Utility is approximated through a hidden grid search instead of VBA UDFs."),
        ("Inputs", "Change province/year, income, or price multipliers in RUN."),
        ("Main units", "Food x = daily kcal / 2000; prices and m are in 2023 real-price terms."),
        ("Caution", "Future years and income points outside sample support are projections/extrapolations."),
    ]
    for r, (k, v) in enumerate(readme_rows, 4):
        ws_readme.cell(r, 1, k).font = Font(bold=True)
        ws_readme.cell(r, 2, v)
        ws_readme.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws_readme["A13"] = "Sheets"
    ws_readme["A13"].font = Font(bold=True, size=12)
    sheet_notes = [
        ("RUN", "Interactive simulation page."),
        ("MAIDADS MODEL", "Equations and units used by the workbook."),
        ("Province Data", "Observed and projected province-year inputs."),
        ("Projection Data", "National projection and feed-grain outputs."),
        ("Elasticities", "Estimated income and price elasticity tables."),
        ("Checks", "Formula checks and diagnostics."),
        ("Utility Solver", "Hidden grid used to solve utility for base and +/- income."),
    ]
    for r, (s, d) in enumerate(sheet_notes, 15):
        ws_readme.cell(r, 1, s).font = Font(bold=True)
        ws_readme.cell(r, 2, d)
        ws_readme.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws_readme.column_dimensions["A"].width = 22
    ws_readme.column_dimensions["B"].width = 78
    for row in range(4, 10):
        ws_readme.row_dimensions[row].height = 24

    # Model sheet
    ws_model.merge_cells("A1:J1")
    ws_model["A1"] = "MAIDADS Equations Used in This Workbook"
    style_title(ws_model["A1"])
    model_lines = [
        ("Demand", "x_i = gamma_i(u) + phi_i(u) * [m - SUM_j p_j gamma_j(u)] / p_i"),
        ("Marginal budget share", "phi_i(u) = [alpha_i + beta_i exp(u)] / [1 + exp(u)]"),
        ("Subsistence", "gamma_i(u) = [delta_i + tau_i exp(omega u)] / [1 + exp(omega u)]"),
        ("Utility", "SUM_i phi_i(u) ln[x_i - gamma_i(u)] - u - kappa = 0"),
        ("Income elasticity", "Computed by central difference using m*(1 +/- step)."),
        ("Price elasticity", "Marshallian formula follows the LES-like MAIDADS price response; Hicksian matrix is in the results files."),
        ("Checks", "Budget shares sum to one; weighted income elasticities sum to one; implicit utility gap should be near zero."),
    ]
    for r, (k, v) in enumerate(model_lines, 4):
        ws_model.cell(r, 1, k).font = Font(bold=True)
        ws_model.cell(r, 2, v)
    ws_model.column_dimensions["A"].width = 24
    ws_model.column_dimensions["B"].width = 120

    # Lists and data
    provinces = sorted(province_data["provincechn"].unique(), key=lambda x: (x != "全国加权", x))
    years = sorted(province_data["year"].unique())
    for i, p in enumerate(provinces, 2):
        ws_lists.cell(i, 1, p)
    ws_lists["A1"] = "Province"
    for i, y in enumerate(years, 2):
        ws_lists.cell(i, 2, y)
    ws_lists["B1"] = "Year"
    ws_lists.sheet_state = "hidden"

    data_last = write_df(ws_data, province_data)
    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = f"A1:{get_column_letter(ws_data.max_column)}{data_last}"
    for col in range(1, ws_data.max_column + 1):
        ws_data.column_dimensions[get_column_letter(col)].width = 16 if col > 6 else 18
    for col in range(7, ws_data.max_column + 1):
        for cell in ws_data.iter_cols(min_col=col, max_col=col, min_row=2, max_row=data_last):
            for c in cell:
                c.number_format = "0.0000"

    # Projection sheet
    ws_proj.merge_cells("A1:H1")
    ws_proj["A1"] = "Projection Outputs"
    style_title(ws_proj["A1"])
    group_pivot = projection_group.pivot_table(
        index="group", columns="year", values="daily_kcal_per_cap_weighted"
    ).reset_index()
    group_pivot.columns = [str(c) for c in group_pivot.columns]
    ws_proj["A3"] = "National weighted daily kcal per capita"
    ws_proj["A3"].font = Font(bold=True, size=12)
    group_end = write_df(ws_proj, group_pivot, 4, 1)
    feed = projection_feed.copy()
    feed["feed_grain_million_ton"] = feed["feed_grain_kg"] / 1e9
    feed_pivot = feed.pivot_table(index="item", columns="year", values="feed_grain_million_ton").reset_index()
    feed_pivot.columns = [str(c) for c in feed_pivot.columns]
    ws_proj["A13"] = "Feed-grain demand, million tons"
    ws_proj["A13"].font = Font(bold=True, size=12)
    feed_end = write_df(ws_proj, feed_pivot, 14, 1)
    style_table(ws_proj, 4, group_end, 1, group_pivot.shape[1])
    style_table(ws_proj, 14, feed_end, 1, feed_pivot.shape[1])
    for col in range(1, 7):
        ws_proj.column_dimensions[get_column_letter(col)].width = 18

    chart = LineChart()
    chart.title = "Daily kcal by group"
    chart.y_axis.title = "kcal/person/day"
    chart.x_axis.title = "Group"
    data_ref = Reference(ws_proj, min_col=2, max_col=group_pivot.shape[1], min_row=4, max_row=group_end)
    cats_ref = Reference(ws_proj, min_col=1, min_row=5, max_row=group_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 15
    ws_proj.add_chart(chart, "G4")

    # Elasticities sheet
    ws_elast.merge_cells("A1:J1")
    ws_elast["A1"] = "Elasticity Tables"
    style_title(ws_elast["A1"])
    inc = income_elast[income_elast["group"].isin(GROUPS + ["all_food", "animal_food", "plant_food"])].copy()
    inc = inc[["income", "group", "eta", "budget_share", "support_flag"]]
    ws_elast["A3"] = "Income elasticity grid"
    ws_elast["A3"].font = Font(bold=True, size=12)
    inc_end = write_df(ws_elast, inc, 4, 1)
    own = price_elast[price_elast["is_own_price"].astype(bool)].copy()
    own = own[["income", "demand_group", "elasticity", "support_flag"]]
    own_start = inc_end + 3
    ws_elast.cell(own_start, 1, "Marshallian own-price elasticities").font = Font(bold=True, size=12)
    own_end = write_df(ws_elast, own, own_start + 1, 1)
    style_table(ws_elast, 4, inc_end, 1, 5)
    style_table(ws_elast, own_start + 1, own_end, 1, 4)
    ws_elast.freeze_panes = "A4"
    ws_elast.auto_filter.ref = f"A4:E{inc_end}"
    for col in range(1, 8):
        ws_elast.column_dimensions[get_column_letter(col)].width = 18

    # RUN sheet layout
    ws_run.merge_cells("A1:R1")
    ws_run["A1"] = "Provincial MAIDADS Simulator"
    style_title(ws_run["A1"])
    ws_run["B4"] = "Inputs"
    ws_run["B4"].font = Font(bold=True, size=12)
    input_rows = [
        ("Province", "全国加权"),
        ("Year", 2023),
        ("Source", '=INDEX(\'Province Data\'!$B$2:$B$%d,MATCH($C$15,\'Province Data\'!$A$2:$A$%d,0))' % (data_last, data_last)),
        ("Income multiplier", 1.0),
        ("Manual m override", None),
        ("Elasticity step", 0.0001),
    ]
    for idx, (label, value) in enumerate(input_rows, 8):
        ws_run.cell(idx, 2, label).font = Font(bold=True)
        ws_run.cell(idx, 3, value)
    ws_run["B15"] = "Lookup key"
    ws_run["C15"] = '=$C$8&"|"&$C$9'
    ws_run["B16"] = "Budget m"
    ws_run["C16"] = '=IF(ISBLANK($C$12),INDEX(\'Province Data\'!$G$2:$G$%d,MATCH($C$15,\'Province Data\'!$A$2:$A$%d,0))*$C$11,$C$12)' % (data_last, data_last)
    ws_run["B17"] = "Utility u"
    ws_run["C17"] = '=INDEX(\'Utility Solver\'!$A$2:$A$802,MATCH(MIN(\'Utility Solver\'!$V$2:$V$802),\'Utility Solver\'!$V$2:$V$802,0))'
    ws_run["B18"] = "Utility u, m+"
    ws_run["C18"] = '=INDEX(\'Utility Solver\'!$A$2:$A$802,MATCH(MIN(\'Utility Solver\'!$AE$2:$AE$802),\'Utility Solver\'!$AE$2:$AE$802,0))'
    ws_run["B19"] = "Utility u, m-"
    ws_run["C19"] = '=INDEX(\'Utility Solver\'!$A$2:$A$802,MATCH(MIN(\'Utility Solver\'!$AN$2:$AN$802),\'Utility Solver\'!$AN$2:$AN$802,0))'
    ws_run["B20"] = "Implicit utility gap"
    ws_run["C20"] = '=INDEX(\'Utility Solver\'!$U$2:$U$802,MATCH(MIN(\'Utility Solver\'!$V$2:$V$802),\'Utility Solver\'!$V$2:$V$802,0))'
    for cell in ws_run["C8:C20"]:
        cell[0].fill = PatternFill("solid", fgColor="FFF7D6")
    dv_prov = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${len(provinces)+1}", allow_blank=False)
    dv_year = DataValidation(type="list", formula1=f"=Lists!$B$2:$B${len(years)+1}", allow_blank=False)
    ws_run.add_data_validation(dv_prov)
    ws_run.add_data_validation(dv_year)
    dv_prov.add(ws_run["C8"])
    dv_year.add(ws_run["C9"])

    headers = [
        "Food group",
        "alpha",
        "beta",
        "delta",
        "tau",
        "omega",
        "kappa",
        "Base price",
        "Price mult.",
        "Price p",
        "gamma",
        "phi",
        "Demand x",
        "Income elast.",
        "Budget share",
        "Expenditure",
        "Daily kcal",
    ]
    start_row = 24
    for c, h in enumerate(headers, 2):
        ws_run.cell(start_row, c, h)
    style_header(ws_run[start_row])
    price_data_cols = {g: province_data.columns.get_loc(f"p_{g}") + 1 for g in GROUPS}
    for idx, g in enumerate(GROUPS, start_row + 1):
        param = params[params["group"].eq(g)].iloc[0]
        ws_run.cell(idx, 2, GROUP_LABELS[g])
        for c, field in zip(range(3, 9), ["alpha", "beta", "delta", "tau", "omega", "kappa"]):
            ws_run.cell(idx, c, float(param[field]))
        price_col_letter = get_column_letter(price_data_cols[g])
        ws_run.cell(idx, 9, f"=INDEX('Province Data'!${price_col_letter}$2:${price_col_letter}${data_last},MATCH($C$15,'Province Data'!$A$2:$A${data_last},0))")
        ws_run.cell(idx, 10, 1.0)
        ws_run.cell(idx, 11, f"=I{idx}*J{idx}")
        ws_run.cell(idx, 12, f"=(E{idx}+F{idx}*EXP(G{idx}*$C$17))/(1+EXP(G{idx}*$C$17))")
        ws_run.cell(idx, 13, f"=(C{idx}+D{idx}*EXP($C$17))/(1+EXP($C$17))")
        ws_run.cell(idx, 14, f"=L{idx}+M{idx}*($C$16-SUMPRODUCT($K$25:$K$30,$L$25:$L$30))/K{idx}")
        ws_run.cell(idx, 15, f'=IFERROR((LN(U{idx})-LN(X{idx}))/(LN($C$16*(1+$C$13))-LN($C$16*(1-$C$13))),"")')
        ws_run.cell(idx, 16, f"=N{idx}*K{idx}/$C$16")
        ws_run.cell(idx, 17, f"=N{idx}*K{idx}")
        ws_run.cell(idx, 18, "" if g == "nonfood" else f"=N{idx}*2000")
        ws_run.cell(idx, 19, f"=(E{idx}+F{idx}*EXP(G{idx}*$C$18))/(1+EXP(G{idx}*$C$18))")
        ws_run.cell(idx, 20, f"=(C{idx}+D{idx}*EXP($C$18))/(1+EXP($C$18))")
        ws_run.cell(idx, 21, f"=S{idx}+T{idx}*($C$16*(1+$C$13)-SUMPRODUCT($K$25:$K$30,$S$25:$S$30))/K{idx}")
        ws_run.cell(idx, 22, f"=(E{idx}+F{idx}*EXP(G{idx}*$C$19))/(1+EXP(G{idx}*$C$19))")
        ws_run.cell(idx, 23, f"=(C{idx}+D{idx}*EXP($C$19))/(1+EXP($C$19))")
        ws_run.cell(idx, 24, f"=V{idx}+W{idx}*($C$16*(1-$C$13)-SUMPRODUCT($K$25:$K$30,$V$25:$V$30))/K{idx}")
    total_row = start_row + 7
    ws_run.cell(total_row, 2, "Total / checks").font = Font(bold=True)
    ws_run.cell(total_row, 13, "=SUM(M25:M30)")
    ws_run.cell(total_row, 15, "=SUMPRODUCT(O25:O30,P25:P30)")
    ws_run.cell(total_row, 16, "=SUM(P25:P30)")
    ws_run.cell(total_row, 17, "=SUM(Q25:Q30)")
    ws_run.cell(total_row, 18, "=SUM(R25:R29)")
    style_table(ws_run, start_row, total_row, 2, 18)
    for col in range(19, 25):
        ws_run.column_dimensions[get_column_letter(col)].hidden = True

    matrix_row = 35
    ws_run.cell(matrix_row, 2, "Marshallian price elasticities").font = Font(bold=True, size=12)
    header_row = matrix_row + 1
    ws_run.cell(header_row, 2, "Demand \\ Price")
    for j, g in enumerate(GROUPS, 3):
        ws_run.cell(header_row, j, GROUP_LABELS[g])
    style_header(ws_run[header_row])
    for i, g_i in enumerate(GROUPS, header_row + 1):
        ws_run.cell(i, 2, GROUP_LABELS[g_i])
        demand_row = start_row + 1 + GROUPS.index(g_i)
        for j, g_j in enumerate(GROUPS, 3):
            price_row = start_row + 1 + GROUPS.index(g_j)
            own = "1" if g_i == g_j else "0"
            ws_run.cell(
                i,
                j,
                f"=($M${price_row}*(($C$16-SUMPRODUCT($K$25:$K$30,$L$25:$L$30))/($P${price_row}*$C$16)))*($M${demand_row}-{own})-($P${demand_row}*$O${price_row})",
            )
    style_table(ws_run, header_row, header_row + len(GROUPS), 2, 2 + len(GROUPS))

    hicks_row = header_row + len(GROUPS) + 3
    ws_run.cell(hicks_row, 2, "Hicksian price elasticities").font = Font(bold=True, size=12)
    ws_run.cell(hicks_row + 1, 2, "Demand \\ Price")
    for j, g in enumerate(GROUPS, 3):
        ws_run.cell(hicks_row + 1, j, GROUP_LABELS[g])
    style_header(ws_run[hicks_row + 1])
    for i, g_i in enumerate(GROUPS, hicks_row + 2):
        ws_run.cell(i, 2, GROUP_LABELS[g_i])
        demand_idx = start_row + 1 + GROUPS.index(g_i)
        source_mar_row = header_row + 1 + GROUPS.index(g_i)
        for j, g_j in enumerate(GROUPS, 3):
            price_idx = start_row + 1 + GROUPS.index(g_j)
            mar_cell = f"{get_column_letter(j)}{source_mar_row}"
            ws_run.cell(i, j, f"={mar_cell}+$O${demand_idx}*$P${price_idx}")
    style_table(ws_run, hicks_row + 1, hicks_row + 1 + len(GROUPS), 2, 2 + len(GROUPS))

    for col, width in {
        "A": 3,
        "B": 28,
        "C": 16,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 14,
        "J": 12,
        "K": 14,
        "L": 12,
        "M": 12,
        "N": 14,
        "O": 14,
        "P": 14,
        "Q": 14,
        "R": 14,
    }.items():
        ws_run.column_dimensions[col].width = width
    ws_run.freeze_panes = "B24"

    # Utility Solver sheet
    solver_headers = (
        ["u"]
        + [f"gamma_{g}" for g in GROUPS]
        + [f"phi_{g}" for g in GROUPS]
        + ["disc_base"]
        + [f"qdisc_base_{g}" for g in GROUPS]
        + ["gap_base", "abs_gap_base", "disc_plus"]
        + [f"qdisc_plus_{g}" for g in GROUPS]
        + ["gap_plus", "abs_gap_plus", "disc_minus"]
        + [f"qdisc_minus_{g}" for g in GROUPS]
        + ["gap_minus", "abs_gap_minus"]
    )
    for c, h in enumerate(solver_headers, 1):
        ws_solver.cell(1, c, h)
    style_header(ws_solver[1])
    u_values = np.round(np.linspace(-20, 20, 801), 6)
    for r, u in enumerate(u_values, 2):
        ws_solver.cell(r, 1, float(u))
        for j, g in enumerate(GROUPS):
            run_row = start_row + 1 + j
            gamma_col = 2 + j
            phi_col = 8 + j
            ws_solver.cell(r, gamma_col, f"=(RUN!$E${run_row}+RUN!$F${run_row}*EXP(RUN!$G${run_row}*$A{r}))/(1+EXP(RUN!$G${run_row}*$A{r}))")
            ws_solver.cell(r, phi_col, f"=(RUN!$C${run_row}+RUN!$D${run_row}*EXP($A{r}))/(1+EXP($A{r}))")
        ws_solver.cell(r, 14, f"=RUN!$C$16-SUMPRODUCT(RUN!$K$25:$K$30,B{r}:G{r})")
        for j in range(6):
            run_row = start_row + 1 + j
            ws_solver.cell(r, 15 + j, f"=IF($N{r}<=0,NA(),{get_column_letter(8+j)}{r}*$N{r}/RUN!$K${run_row})")
        ws_solver.cell(r, 21, f'=IF($N{r}<=0,1E+99,SUMPRODUCT(H{r}:M{r},LN(O{r}:T{r}))-$A{r}-RUN!$H$25)')
        ws_solver.cell(r, 22, f"=ABS(U{r})")
        ws_solver.cell(r, 23, f"=RUN!$C$16*(1+RUN!$C$13)-SUMPRODUCT(RUN!$K$25:$K$30,B{r}:G{r})")
        for j in range(6):
            run_row = start_row + 1 + j
            ws_solver.cell(r, 24 + j, f"=IF($W{r}<=0,NA(),{get_column_letter(8+j)}{r}*$W{r}/RUN!$K${run_row})")
        ws_solver.cell(r, 30, f'=IF($W{r}<=0,1E+99,SUMPRODUCT(H{r}:M{r},LN(X{r}:AC{r}))-$A{r}-RUN!$H$25)')
        ws_solver.cell(r, 31, f"=ABS(AD{r})")
        ws_solver.cell(r, 32, f"=RUN!$C$16*(1-RUN!$C$13)-SUMPRODUCT(RUN!$K$25:$K$30,B{r}:G{r})")
        for j in range(6):
            run_row = start_row + 1 + j
            ws_solver.cell(r, 33 + j, f"=IF($AF{r}<=0,NA(),{get_column_letter(8+j)}{r}*$AF{r}/RUN!$K${run_row})")
        ws_solver.cell(r, 39, f'=IF($AF{r}<=0,1E+99,SUMPRODUCT(H{r}:M{r},LN(AG{r}:AL{r}))-$A{r}-RUN!$H$25)')
        ws_solver.cell(r, 40, f"=ABS(AM{r})")
    ws_solver.freeze_panes = "A2"
    ws_solver.sheet_state = "hidden"

    # Checks sheet
    ws_checks.merge_cells("A1:H1")
    ws_checks["A1"] = "Workbook Checks"
    style_title(ws_checks["A1"])
    checks = [
        ("Phi sum", "=RUN!M31", "Should equal 1."),
        ("Budget share sum", "=RUN!P31", "Should equal 1."),
        ("Weighted income elasticity", "=RUN!O31", "Should equal 1."),
        ("Implicit utility gap", "=RUN!C20", "Should be close to 0; grid approximation tolerance depends on utility grid step."),
        ("Budget identity", "=RUN!Q31-RUN!C16", "Should be close to 0."),
        ("Main MAIDADS nll", "=INDEX(model_comparison_nll,MATCH(\"MAIDADS_sat\",model_comparison_model,0))", "Reference only; formula names are not defined in this workbook."),
    ]
    ws_checks.append(["Check", "Value", "Interpretation"])
    style_header(ws_checks[2])
    for r, (name, formula, note) in enumerate(checks, 3):
        ws_checks.cell(r, 1, name)
        ws_checks.cell(r, 2, formula)
        ws_checks.cell(r, 3, note)
    ws_checks["A12"] = "Consistency tests from results file"
    ws_checks["A12"].font = Font(bold=True, size=12)
    cons_end = write_df(ws_checks, consistency, 13, 1)
    style_table(ws_checks, 2, 8, 1, 3)
    style_table(ws_checks, 13, cons_end, 1, consistency.shape[1])
    ws_checks.column_dimensions["A"].width = 34
    ws_checks.column_dimensions["B"].width = 18
    ws_checks.column_dimensions["C"].width = 90
    for col in range(4, consistency.shape[1] + 1):
        ws_checks.column_dimensions[get_column_letter(col)].width = 18

    # Remove unsupported named-formula check row value.
    ws_checks["B8"] = "See model_comparison.csv"

    # Formatting across RUN.
    for row in ws_run.iter_rows(min_row=25, max_row=31, min_col=3, max_col=18):
        for cell in row:
            cell.number_format = "0.0000"
    for row in ws_run.iter_rows(min_row=37, max_row=55, min_col=3, max_col=8):
        for cell in row:
            cell.number_format = "0.0000"
    ws_run["C16"].number_format = "#,##0.00"
    ws_run["C17"].number_format = "0.0000"
    ws_run["C18"].number_format = "0.0000"
    ws_run["C19"].number_format = "0.0000"
    ws_run["C20"].number_format = "0.000000"

    for ws in [ws_proj, ws_elast, ws_data, ws_checks]:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0000"

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUTPUT)
    # Reopen once to ensure the package is readable.
    check = load_workbook(OUTPUT, data_only=False, read_only=True)
    required = {"Read Me", "MAIDADS MODEL", "RUN", "Province Data", "Projection Data", "Elasticities", "Checks"}
    missing = required.difference(check.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook missing sheets: {sorted(missing)}")
    check.close()
    print(OUTPUT)


if __name__ == "__main__":
    main()
