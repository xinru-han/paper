#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Paper 8 script 01 (= plan 80a): parse prefecture daily weather xlsx (1973-2024)
into a single long csv.gz.

Input : RAW/地级市天气数据/城市/{YYYY}年各城市{气温|平均降水量}日度数据.xlsx
        temperature files: 平均气温/日期/最高气温/最低气温/省/省代码/市/市代码
        precipitation files: same layout with 平均降水量 as first column
Output: data/interim/city_day_weather_{YYYY}.csv (per year, then combined
        by 01b in R). Values rounded to 2dp to keep files small.
"""
import csv
import glob
import os
import re
import sys
from openpyxl import load_workbook

RAW = "/root/data/数据/央视数据/地级市天气数据/城市"
OUT = "/root/data/Paper/央视数据/paper8-hot/data/interim/weather_years"
os.makedirs(OUT, exist_ok=True)

YEARS = sys.argv[1:] if len(sys.argv) > 1 else [str(y) for y in range(1973, 2025)]


def read_file(path, value_cols):
    """Return dict keyed by (city_code, date) -> tuple of values + meta."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}
    need = value_cols + ["日期", "省", "省代码", "市", "市代码"]
    for c in need:
        if c not in idx:
            raise KeyError(f"{os.path.basename(path)} missing column {c}: {header}")
    out = {}
    for r in rows:
        if r is None or r[idx["日期"]] is None:
            continue
        d = r[idx["日期"]]
        date = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10]
        key = (r[idx["市代码"]], date)
        vals = []
        for c in value_cols:
            v = r[idx[c]]
            vals.append(round(float(v), 2) if isinstance(v, (int, float)) else "")
        out[key] = (vals, r[idx["省"]], r[idx["省代码"]], r[idx["市"]])
    wb.close()
    return out


for y in YEARS:
    dest = os.path.join(OUT, f"city_day_weather_{y}.csv")
    if os.path.exists(dest) and os.path.getsize(dest) > 1000:
        continue
    ft = glob.glob(os.path.join(RAW, f"{y}年各城市气温日度数据.xlsx"))
    fp = glob.glob(os.path.join(RAW, f"{y}年各城市平均降水量日度数据.xlsx"))
    if not ft:
        print(f"[{y}] no temperature file, skip", flush=True)
        continue
    temp = read_file(ft[0], ["平均气温", "最高气温", "最低气温"])
    prec = read_file(fp[0], ["降水量"]) if fp else {}
    with open(dest + ".tmp", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["city_code", "date", "tavg", "tmax", "tmin", "precip",
                    "province", "province_code", "city"])
        for (cc, date), (vals, prov, pcode, city) in temp.items():
            pv = prec.get((cc, date))
            w.writerow([cc, date] + vals + [pv[0][0] if pv else ""] +
                       [prov, pcode, city])
    os.replace(dest + ".tmp", dest)
    print(f"[{y}] {len(temp)} rows -> {os.path.basename(dest)}", flush=True)

print("done", flush=True)
