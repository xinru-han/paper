"""Emulate the subset of GDXXRW behaviour used by CASM.

The GAMS code loads data from Excel workbooks via GDXXRW with an INDEX
sheet.  Every entry of the INDEX sheet is a (type, symbol, range, rdim,
cdim) tuple.  This module reads the same ranges directly with openpyxl and
returns plain Python dictionaries keyed by label tuples, so the rest of
the port can look values up exactly the way GAMS does.

GDXXRW conventions replicated here (options SE=0 used by CASM):
* PAR with rdim=r, cdim=c anchored at cell (R, C): column labels occupy
  rows R .. R+c-1 starting at column C+r; row labels occupy columns
  C .. C+r-1 starting at row R+c.  Values live in the intersection.
* Scanning stops at the first row/column whose label cells are all empty
  (SkipEmpty = 0).
* For multi-dimensional headers, outer-dimension labels carry forward to
  the right/downward when blank.
* Labels are case-insensitive; we normalise to upper case and strip
  blanks.  Empty cells mean "value 0" (GAMS does not store zeros).
* DSET reads the unique labels of a single column starting at the anchor
  cell itself.  SET (rdim=n) reads n label columns per row, one tuple per
  row, until an empty row.
"""

from __future__ import annotations

import re
from openpyxl.utils import column_index_from_string


def _norm(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    if s == "":
        return None
    # years may come back as floats (2023.0) from Excel
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def parse_ref(ref):
    """'RICE0!E6' -> (sheet, row, col) with row/col 1-based."""
    sheet, cell = ref.split("!")
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", cell.strip())
    col = column_index_from_string(m.group(1).upper())
    row = int(m.group(2))
    return sheet.strip().strip("'"), row, col


class Sheet:
    """Random access wrapper over a worksheet's used range."""

    def __init__(self, ws):
        self.cells = {}
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for c, v in enumerate(row, start=1):
                if v is not None and not (isinstance(v, str) and v.strip() == ""):
                    self.cells[(r, c)] = v
        self.max_row = max((k[0] for k in self.cells), default=0)
        self.max_col = max((k[1] for k in self.cells), default=0)

    def get(self, r, c):
        return self.cells.get((r, c))


class Workbook:
    def __init__(self, path):
        import openpyxl

        # data_only=True: use cached formula results (the INDEX sheets of the
        # CASM workbooks contain CONCATENATE formulas for range references)
        self._wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self._sheets = {}
        # worksheet titles are matched case-insensitively like GDXXRW does
        self._titlemap = {name.strip().upper(): name for name in self._wb.sheetnames}

    def sheet(self, name):
        key = name.strip().upper()
        if key not in self._sheets:
            self._sheets[key] = Sheet(self._wb[self._titlemap[key]])
        return self._sheets[key]


def read_dset(wb: Workbook, ref):
    """Domain set: one column of labels starting at the anchor cell."""
    sheet, r0, c0 = parse_ref(ref)
    sh = wb.sheet(sheet)
    labels = []
    r = r0
    while True:
        v = _norm(sh.get(r, c0))
        if v is None:
            break
        if v not in labels:
            labels.append(v)
        r += 1
    return labels


def read_set(wb: Workbook, ref, rdim):
    """Multi-column set: one tuple per row until an empty row."""
    sheet, r0, c0 = parse_ref(ref)
    sh = wb.sheet(sheet)
    out = []
    r = r0
    while True:
        tup = tuple(_norm(sh.get(r, c0 + d)) for d in range(rdim))
        if all(v is None for v in tup):
            break
        if all(v is not None for v in tup) and tup not in out:
            out.append(tup)
        r += 1
    return out


def read_par(wb: Workbook, ref, rdim, cdim):
    """Parameter block -> dict[label-tuple] = float."""
    sheet, r0, c0 = parse_ref(ref)
    sh = wb.sheet(sheet)
    rdim = int(rdim or 0)
    cdim = int(cdim or 0)

    # ---- column headers ----------------------------------------------
    col_keys = []  # (col_index, tuple-of-labels)
    if cdim:
        c = c0 + rdim
        carry = [None] * cdim
        while True:
            raw = [_norm(sh.get(r0 + d, c)) for d in range(cdim)]
            if all(v is None for v in raw):
                break
            for d in range(cdim - 1):  # outer dims carry forward
                if raw[d] is None:
                    raw[d] = carry[d]
                else:
                    carry[d] = raw[d]
            if raw[cdim - 1] is None:
                break
            col_keys.append((c, tuple(raw)))
            c += 1

    # ---- row headers + data ------------------------------------------
    out = {}
    r = r0 + cdim
    carry = [None] * rdim
    if rdim == 0:
        for c, ck in col_keys:
            v = sh.get(r, c)
            if isinstance(v, (int, float)):
                out[ck] = float(v)
        return out
    while True:
        raw = [_norm(sh.get(r, c0 + d)) for d in range(rdim)]
        if all(v is None for v in raw):
            break
        for d in range(rdim - 1):
            if raw[d] is None:
                raw[d] = carry[d]
            else:
                carry[d] = raw[d]
        if raw[rdim - 1] is None:
            break
        rk = tuple(raw)
        if cdim:
            for c, ck in col_keys:
                v = sh.get(r, c)
                if isinstance(v, (int, float)):
                    out[rk + ck] = float(v)
        else:  # single value column right of the row labels
            v = sh.get(r, c0 + rdim)
            if isinstance(v, (int, float)):
                out[rk] = float(v)
        r += 1
    return out


def read_index(wb: Workbook, ref):
    """Parse an INDEX sheet starting at `ref` (e.g. 'INDEX!A1').

    Returns list of (kind, symbol, range, rdim, cdim)."""
    sheet, r0, c0 = parse_ref(ref)
    sh = wb.sheet(sheet)
    entries = []
    for r in range(r0 + 1, sh.max_row + 1):
        kind = _norm(sh.get(r, c0))
        if kind is None:
            continue
        if kind not in ("PAR", "SET", "DSET"):
            continue
        sym = _norm(sh.get(r, c0 + 1))
        rng = sh.get(r, c0 + 2)
        rdim = sh.get(r, c0 + 3)
        cdim = sh.get(r, c0 + 4)
        entries.append((kind, sym, str(rng).strip(), rdim, cdim))
    return entries


def load_symbols(path, index_ref, wanted=None):
    """Load all (or `wanted`) symbols of a workbook according to its INDEX."""
    wb = Workbook(path)
    out = {}
    for kind, sym, rng, rdim, cdim in read_index(wb, index_ref):
        if wanted is not None and sym not in wanted:
            continue
        if kind == "DSET":
            out[sym] = read_dset(wb, rng)
        elif kind == "SET":
            out[sym] = read_set(wb, rng, int(rdim or 1))
        else:
            out[sym] = read_par(wb, rng, rdim, cdim)
    return out
